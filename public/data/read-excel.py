import json
from openpyxl import load_workbook

excel_file = r"E:\xampp\htdocs\safety-competition-questions\public\data\data-excel.xlsx"

output_file = r"E:\xampp\htdocs\safety-competition-questions\public\data\data-question-answer.json"

wb = load_workbook(excel_file, data_only=True)

print("Sheets:", wb.sheetnames)

ws = wb["Sheet1"]

data = []

# for row in ws.iter_rows(min_row=2, values_only=True):
#     if row[0] is None:
#         continue

#     data.append({
#         "serial_no": str(row[0]).strip(),
#         "question": str(row[1]).strip() if row[1] else "",
#         "answer": str(row[2]).strip() if row[2] else ""
#     })

for row in ws.iter_rows(min_row=2, values_only=True):

    serial_no = row[0]
    question = row[1]
    answer = row[2]

    if serial_no is None:
        continue

    if question is None or answer is None:
        continue

    data.append({
        "serial_no": str(serial_no).strip(),
        "question": str(question).strip(),
        "answer": str(answer).strip()
    })

with open(output_file, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=4)

print(f"Saved {len(data)} records")
print(f"Output: {output_file}")